import sqlite3
from datetime import datetime, timedelta
import os

class Database:
    def __init__(self, db_path='database/licoreria.db'):
        self.db_path = db_path
        self.inicializar_base_datos()
    
    def get_connection(self):
        """Crear conexión a la base de datos"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn
    
    def inicializar_base_datos(self):
        """Crear todas las tablas necesarias"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Tabla: productos
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                descripcion TEXT,
                imagen TEXT,
                precio_compra REAL NOT NULL,
                precio_venta REAL NOT NULL,
                unidad TEXT NOT NULL,
                categoria TEXT NOT NULL,
                stock INTEGER NOT NULL DEFAULT 0,
                stock_minimo INTEGER NOT NULL DEFAULT 5,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_modificacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Tabla: ventas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS ventas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                total REAL NOT NULL,
                metodo_pago TEXT NOT NULL,
                monto_efectivo REAL DEFAULT 0,
                monto_qr REAL DEFAULT 0,
                cliente_nombre TEXT,
                cliente_telefono TEXT,
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                estado TEXT DEFAULT 'completada'
            )
        ''')
        
        # Tabla: detalle_ventas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS detalle_ventas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                venta_id INTEGER NOT NULL,
                producto_id INTEGER NOT NULL,
                producto_nombre TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                precio_unitario REAL NOT NULL,
                subtotal REAL NOT NULL,
                FOREIGN KEY (venta_id) REFERENCES ventas(id),
                FOREIGN KEY (producto_id) REFERENCES productos(id)
            )
        ''')
        
        # Tabla: compras
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS compras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tipo TEXT NOT NULL,
                descripcion TEXT,
                monto REAL NOT NULL,
                proveedor TEXT,
                metodo_pago TEXT NOT NULL,
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Tabla: detalle_compras
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS detalle_compras (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                compra_id INTEGER NOT NULL,
                producto_id INTEGER,
                producto_nombre TEXT,
                cantidad INTEGER,
                precio_unitario REAL,
                subtotal REAL,
                FOREIGN KEY (compra_id) REFERENCES compras(id),
                FOREIGN KEY (producto_id) REFERENCES productos(id)
            )
        ''')
        
        # Tabla: creditos
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS creditos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                venta_id INTEGER NOT NULL,
                cliente_nombre TEXT NOT NULL,
                cliente_telefono TEXT,
                monto_total REAL NOT NULL,
                monto_pagado REAL DEFAULT 0,
                saldo_pendiente REAL NOT NULL,
                estado TEXT DEFAULT 'pendiente',
                fecha_credito TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_ultimo_pago TIMESTAMP,
                FOREIGN KEY (venta_id) REFERENCES ventas(id)
            )
        ''')
        
        # Tabla: pagos_creditos
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pagos_creditos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                credito_id INTEGER NOT NULL,
                monto REAL NOT NULL,
                metodo_pago TEXT NOT NULL,
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (credito_id) REFERENCES creditos(id)
            )
        ''')
        
        # Tabla: caja
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS caja (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha_apertura TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                fecha_cierre TIMESTAMP,
                monto_inicial REAL NOT NULL DEFAULT 0,
                total_efectivo REAL DEFAULT 0,
                total_qr REAL DEFAULT 0,
                total_credito REAL DEFAULT 0,
                total_ingresos REAL DEFAULT 0,
                total_egresos REAL DEFAULT 0,
                efectivo_esperado REAL DEFAULT 0,
                efectivo_contado REAL DEFAULT 0,
                diferencia REAL DEFAULT 0,
                estado TEXT DEFAULT 'abierta',
                usuario TEXT
            )
        ''')
        
        # Tabla: movimientos_caja
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS movimientos_caja (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                caja_id INTEGER NOT NULL,
                tipo TEXT NOT NULL,
                concepto TEXT NOT NULL,
                monto REAL NOT NULL,
                metodo_pago TEXT,
                referencia_id INTEGER,
                referencia_tipo TEXT,
                fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (caja_id) REFERENCES caja(id)
            )
        ''')
        
        conn.commit()
        conn.close()
        print('✅ Base de datos inicializada correctamente')
    
    # ========== FUNCIONES PARA PRODUCTOS ==========
    
    def crear_producto(self, nombre, descripcion, imagen, precio_compra, precio_venta, 
                      unidad, categoria, stock, stock_minimo):
        """Crear un nuevo producto"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO productos (
                    nombre, descripcion, imagen, precio_compra, precio_venta,
                    unidad, categoria, stock, stock_minimo
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (nombre, descripcion, imagen, precio_compra, precio_venta, 
                  unidad, categoria, stock, stock_minimo))
            
            producto_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            print(f'✅ Producto creado: {nombre} (ID: {producto_id})')
            return producto_id
        except Exception as e:
            print(f'❌ Error al crear producto: {e}')
            return None
    
    def obtener_productos(self):
        """Obtener todos los productos"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM productos ORDER BY nombre ASC')
            rows = cursor.fetchall()
            
            productos = []
            for row in rows:
                productos.append(dict(row))
            
            conn.close()
            return productos
        except Exception as e:
            print(f'❌ Error al obtener productos: {e}')
            return []
    
    def obtener_producto(self, producto_id):
        """Obtener un producto por ID"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM productos WHERE id = ?', (producto_id,))
            row = cursor.fetchone()
            
            conn.close()
            
            if row:
                return dict(row)
            return None
        except Exception as e:
            print(f'❌ Error al obtener producto: {e}')
            return None
    
    def actualizar_producto(self, id, nombre, descripcion, imagen, precio_compra, 
                          precio_venta, unidad, categoria, stock, stock_minimo):
        """Actualizar un producto existente"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                UPDATE productos SET
                    nombre = ?,
                    descripcion = ?,
                    imagen = ?,
                    precio_compra = ?,
                    precio_venta = ?,
                    unidad = ?,
                    categoria = ?,
                    stock = ?,
                    stock_minimo = ?,
                    fecha_modificacion = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (nombre, descripcion, imagen, precio_compra, precio_venta, 
                  unidad, categoria, stock, stock_minimo, id))
            
            conn.commit()
            conn.close()
            
            print(f'✅ Producto actualizado: {nombre} (ID: {id})')
            return True
        except Exception as e:
            print(f'❌ Error al actualizar producto: {e}')
            return False
    
    def eliminar_producto(self, producto_id):
        """Eliminar un producto"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('DELETE FROM productos WHERE id = ?', (producto_id,))
            
            conn.commit()
            conn.close()
            
            print(f'✅ Producto eliminado (ID: {producto_id})')
            return True
        except Exception as e:
            print(f'❌ Error al eliminar producto: {e}')
            return False
    
    def actualizar_stock(self, producto_id, cantidad, operacion='sumar'):
        """Actualizar stock de un producto"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            if operacion == 'sumar':
                cursor.execute('''
                    UPDATE productos 
                    SET stock = stock + ?,
                        fecha_modificacion = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (cantidad, producto_id))
            else:  # restar
                cursor.execute('''
                    UPDATE productos 
                    SET stock = stock - ?,
                        fecha_modificacion = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (cantidad, producto_id))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f'❌ Error al actualizar stock: {e}')
            return False
    
    def exportar_productos_excel(self, productos):
        """Exportar productos a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Productos"
            
            headers = ['ID', 'Nombre', 'Descripción', 'Categoría', 'Unidad', 
                      'Precio Compra', 'Precio Venta', 'Stock', 'Stock Mínimo', 
                      'Fecha Creación']
            
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_idx, producto in enumerate(productos, 2):
                ws.cell(row=row_idx, column=1, value=producto['id'])
                ws.cell(row=row_idx, column=2, value=producto['nombre'])
                ws.cell(row=row_idx, column=3, value=producto.get('descripcion', ''))
                ws.cell(row=row_idx, column=4, value=producto['categoria'])
                ws.cell(row=row_idx, column=5, value=producto['unidad'])
                ws.cell(row=row_idx, column=6, value=producto['precio_compra'])
                ws.cell(row=row_idx, column=7, value=producto['precio_venta'])
                ws.cell(row=row_idx, column=8, value=producto['stock'])
                ws.cell(row=row_idx, column=9, value=producto['stock_minimo'])
                ws.cell(row=row_idx, column=10, value=producto['fecha_creacion'])
                
                if producto['stock'] <= producto['stock_minimo']:
                    for col in range(1, 11):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                        )
            
            column_widths = [8, 30, 40, 15, 12, 15, 15, 10, 15, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            os.makedirs('exports', exist_ok=True)
            filepath = f'exports/productos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filepath)
            
            print(f'✅ Archivo Excel generado: {filepath}')
            return filepath
        except Exception as e:
            print(f'❌ Error al exportar a Excel: {e}')
            return None
    
    # ========== FUNCIONES PARA VENTAS ==========

    ## 2️⃣ FUNCIÓN crear_venta - CORREGIDA (para que registre en caja):

    # ============================================
# CORRECCIÓN PARA database.py
# ============================================
# INSTRUCCIONES:
# 1. Busca la línea "# ========== FUNCIONES PARA VENTAS =========="
# 2. REEMPLAZA todo el código desde ahí hasta donde comienza crear_detalle_venta
# 3. Con el código de abajo
# ============================================

    # ========== FUNCIONES PARA VENTAS ==========

    def crear_venta(self, total, metodo_pago, monto_efectivo=0, monto_qr=0, 
                    cliente_nombre=None, cliente_telefono=None):
        """Crear una nueva venta Y registrar en caja"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Insertar venta
            cursor.execute('''
                INSERT INTO ventas 
                (total, metodo_pago, monto_efectivo, monto_qr, cliente_nombre, cliente_telefono)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                float(total),
                metodo_pago,
                float(monto_efectivo),
                float(monto_qr),
                cliente_nombre,
                cliente_telefono
            ))
            
            venta_id = cursor.lastrowid
            
            # IMPORTANTE: Registrar en caja (excepto crédito)
            if metodo_pago != 'credito':
                cursor.execute('SELECT id FROM caja WHERE estado = "abierta"')
                caja_abierta = cursor.fetchone()
                
                if caja_abierta:
                    # Para ventas EFECTIVO
                    if metodo_pago == 'efectivo':
                        cursor.execute('''
                            INSERT INTO movimientos_caja 
                            (caja_id, tipo, concepto, monto, metodo_pago, referencia_id, referencia_tipo)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            caja_abierta['id'],
                            'ingreso',
                            f"Venta #{venta_id}",
                            float(total),
                            'efectivo',
                            venta_id,
                            'venta'
                        ))
                    
                    # Para ventas QR
                    elif metodo_pago == 'qr':
                        cursor.execute('''
                            INSERT INTO movimientos_caja 
                            (caja_id, tipo, concepto, monto, metodo_pago, referencia_id, referencia_tipo)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (
                            caja_abierta['id'],
                            'ingreso',
                            f"Venta #{venta_id}",
                            float(total),
                            'qr',
                            venta_id,
                            'venta'
                        ))
                    
                    # Para ventas MIXTAS
                    elif metodo_pago == 'mixto':
                        # Registrar parte efectivo
                        if float(monto_efectivo) > 0:
                            cursor.execute('''
                                INSERT INTO movimientos_caja 
                                (caja_id, tipo, concepto, monto, metodo_pago, referencia_id, referencia_tipo)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                caja_abierta['id'],
                                'ingreso',
                                f"Venta #{venta_id} - Efectivo",
                                float(monto_efectivo),
                                'efectivo',
                                venta_id,
                                'venta'
                            ))
                        
                        # Registrar parte QR
                        if float(monto_qr) > 0:
                            cursor.execute('''
                                INSERT INTO movimientos_caja 
                                (caja_id, tipo, concepto, monto, metodo_pago, referencia_id, referencia_tipo)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            ''', (
                                caja_abierta['id'],
                                'ingreso',
                                f"Venta #{venta_id} - QR",
                                float(monto_qr),
                                'qr',
                                venta_id,
                                'venta'
                            ))
            
            conn.commit()
            conn.close()
            
            print(f'✅ Venta creada: #{venta_id} - Bs. {total}')
            return venta_id
            
        except Exception as e:
            print(f'❌ Error al crear venta: {e}')
            return None

    # RESTO DE LAS FUNCIONES CONTINÚAN AQUÍ...
    
    
    def crear_detalle_venta(self, venta_id, producto_id, producto_nombre, 
                           cantidad, precio_unitario, subtotal):
        """Crear detalle de una venta"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO detalle_ventas (
                    venta_id, producto_id, producto_nombre,
                    cantidad, precio_unitario, subtotal
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (venta_id, producto_id, producto_nombre,
                  cantidad, precio_unitario, subtotal))
            
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f'❌ Error al crear detalle de venta: {e}')
            return False
    
    def obtener_ventas(self):
        """Obtener todas las ventas"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM ventas ORDER BY fecha DESC')
            rows = cursor.fetchall()
            
            ventas = []
            for row in rows:
                ventas.append(dict(row))
            
            conn.close()
            return ventas
        except Exception as e:
            print(f'❌ Error al obtener ventas: {e}')
            return []
    
    def obtener_venta(self, venta_id):
        """Obtener una venta por ID"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM ventas WHERE id = ?', (venta_id,))
            row = cursor.fetchone()
            
            conn.close()
            
            if row:
                return dict(row)
            return None
        except Exception as e:
            print(f'❌ Error al obtener venta: {e}')
            return None
    
    def obtener_detalle_venta(self, venta_id):
        """Obtener detalle de una venta"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM detalle_ventas WHERE venta_id = ?', (venta_id,))
            rows = cursor.fetchall()
            
            detalles = []
            for row in rows:
                detalles.append(dict(row))
            
            conn.close()
            return detalles
        except Exception as e:
            print(f'❌ Error al obtener detalle de venta: {e}')
            return []
    
    def obtener_ventas_por_fecha(self, fecha_desde, fecha_hasta):
        """Obtener ventas filtradas por rango de fechas"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                SELECT * FROM ventas 
                WHERE DATE(fecha) BETWEEN ? AND ?
                ORDER BY fecha DESC
            ''', (fecha_desde, fecha_hasta))
            
            rows = cursor.fetchall()
            
            ventas = []
            for row in rows:
                ventas.append(dict(row))
            
            conn.close()
            return ventas
        except Exception as e:
            print(f'❌ Error al obtener ventas por fecha: {e}')
            return []
    
    def exportar_ventas_excel(self, ventas, fecha_desde, fecha_hasta):
        """Exportar ventas a Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Ventas"
            
            headers = ['ID', 'Fecha', 'Hora', 'Método de Pago', 'Cliente', 
                      'Teléfono', 'Efectivo', 'QR', 'Total', 'Estado']
            
            header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            total_general = 0
            for row_idx, venta in enumerate(ventas, 2):
                fecha_obj = datetime.fromisoformat(venta['fecha'].replace('Z', '+00:00'))
                
                ws.cell(row=row_idx, column=1, value=venta['id'])
                ws.cell(row=row_idx, column=2, value=fecha_obj.strftime('%Y-%m-%d'))
                ws.cell(row=row_idx, column=3, value=fecha_obj.strftime('%H:%M:%S'))
                ws.cell(row=row_idx, column=4, value=venta['metodo_pago'].upper())
                ws.cell(row=row_idx, column=5, value=venta.get('cliente_nombre', '-'))
                ws.cell(row=row_idx, column=6, value=venta.get('cliente_telefono', '-'))
                ws.cell(row=row_idx, column=7, value=venta['monto_efectivo'])
                ws.cell(row=row_idx, column=8, value=venta['monto_qr'])
                ws.cell(row=row_idx, column=9, value=venta['total'])
                ws.cell(row=row_idx, column=10, value=venta['estado'].upper())
                
                total_general += venta['total']
            
            # Fila de totales
            row_total = len(ventas) + 2
            ws.cell(row=row_total, column=8, value="TOTAL:").font = Font(bold=True)
            ws.cell(row=row_total, column=9, value=total_general).font = Font(bold=True)
            
            column_widths = [8, 12, 10, 15, 25, 15, 12, 12, 12, 12]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            os.makedirs('exports', exist_ok=True)
            filepath = f'exports/ventas_{fecha_desde}_{fecha_hasta}.xlsx'
            wb.save(filepath)
            
            print(f'✅ Archivo Excel de ventas generado: {filepath}')
            return filepath
        except Exception as e:
            print(f'❌ Error al exportar ventas a Excel: {e}')
            return None
    
    # ========== FUNCIONES PARA CRÉDITOS ==========
    
    def crear_credito(self, venta_id, cliente_nombre, cliente_telefono, monto_total):
        """Crear un crédito"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO creditos (
                    venta_id, cliente_nombre, cliente_telefono,
                    monto_total, saldo_pendiente
                ) VALUES (?, ?, ?, ?, ?)
            ''', (venta_id, cliente_nombre, cliente_telefono, monto_total, monto_total))
            
            credito_id = cursor.lastrowid
            conn.commit()
            conn.close()
            
            print(f'✅ Crédito creado (ID: {credito_id})')
            return credito_id
        except Exception as e:
            print(f'❌ Error al crear crédito: {e}')
            return None

    # ============================================
# CORRECCIÓN PARA database.py - crear_compra
# ============================================
# BUSCA la función crear_compra (debe estar duplicada)
# ELIMINA AMBAS versiones y REEMPLAZA con este código:
# ============================================

    def crear_compra(self, tipo, monto, proveedor, metodo_pago, descripcion=None, fecha=None, observaciones=None):
        """Crear una nueva compra Y registrar en caja"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            if fecha is None:
                fecha = datetime.now()
        
            # Insertar compra
            cursor.execute('''
                INSERT INTO compras (tipo, descripcion, monto, proveedor, metodo_pago, fecha)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (tipo, descripcion, monto, proveedor, metodo_pago, fecha))
        
            compra_id = cursor.lastrowid
            
            # IMPORTANTE: Registrar en caja como EGRESO (las compras salen dinero)
            cursor.execute('SELECT id FROM caja WHERE estado = "abierta"')
            caja_abierta = cursor.fetchone()
            
            if caja_abierta:
                # Determinar concepto según el tipo
                if tipo == 'productos':
                    concepto = f"Compra de productos #{compra_id}"
                elif tipo == 'insumos':
                    concepto = f"Compra de insumos #{compra_id}"
                else:  # gastos
                    concepto = f"Gasto: {descripcion}" if descripcion else f"Gasto #{compra_id}"
                
                # Registrar movimiento de EGRESO en caja
                cursor.execute('''
                    INSERT INTO movimientos_caja 
                    (caja_id, tipo, concepto, monto, metodo_pago, referencia_id, referencia_tipo)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    caja_abierta['id'],
                    'egreso',  # Las compras son EGRESOS (sale dinero)
                    concepto,
                    float(monto),
                    metodo_pago,
                    compra_id,
                    'compra'
                ))
            
            conn.commit()
            conn.close()
            
            print(f'✅ Compra creada: #{compra_id} - {tipo} - Bs. {monto}')
            return compra_id
            
        except Exception as e:
            print(f"❌ Error al crear compra: {e}")
            return None

    def crear_detalle_compra(self, compra_id, producto_id, producto_nombre, cantidad, precio_unitario, subtotal):
        """Crear detalle de compra (para compra de productos)"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            cursor.execute('''
                INSERT INTO detalle_compras 
                (compra_id, producto_id, producto_nombre, cantidad, precio_unitario, subtotal)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (compra_id, producto_id, producto_nombre, cantidad, precio_unitario, subtotal))
        
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error al crear detalle de compra: {e}")
            return False

    def obtener_compras(self, tipo=None, desde=None, hasta=None, buscar=None, limite=100):
        """Obtener compras con filtros opcionales"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            query = "SELECT * FROM compras WHERE 1=1"
            params = []
        
            if tipo:
                query += " AND tipo = ?"
                params.append(tipo)
        
            if desde:
                query += " AND DATE(fecha) >= ?"
                params.append(desde)
        
            if hasta:
                query += " AND DATE(fecha) <= ?"
                params.append(hasta)
        
            if buscar:
                query += " AND (proveedor LIKE ? OR descripcion LIKE ?)"
                params.append(f'%{buscar}%')
                params.append(f'%{buscar}%')
        
            query += " ORDER BY fecha DESC LIMIT ?"
            params.append(limite)
        
            cursor.execute(query, params)
            compras = []
        
            for row in cursor.fetchall():
                compras.append({
                    'id': row['id'],
                    'tipo': row['tipo'],
                    'descripcion': row['descripcion'],
                    'monto': row['monto'],
                    'proveedor': row['proveedor'],
                    'metodo_pago': row['metodo_pago'],
                    'fecha': row['fecha']
                })
        
            conn.close()
            return compras
        except Exception as e:
            print(f"Error al obtener compras: {e}")
            return []

    

    def crear_detalle_compra(self, compra_id, producto_id, producto_nombre, cantidad, precio_unitario, subtotal):
        """Crear detalle de compra (para compra de productos)"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            cursor.execute('''
                INSERT INTO detalle_compras 
                (compra_id, producto_id, producto_nombre, cantidad, precio_unitario, subtotal)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (compra_id, producto_id, producto_nombre, cantidad, precio_unitario, subtotal))
        
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error al crear detalle de compra: {e}")
            return False

    def obtener_compras(self, tipo=None, desde=None, hasta=None, buscar=None, limite=100):
        """Obtener compras con filtros opcionales"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            query = "SELECT * FROM compras WHERE 1=1"
            params = []
        
            if tipo:
                query += " AND tipo = ?"
                params.append(tipo)
        
            if desde:
                query += " AND DATE(fecha) >= ?"
                params.append(desde)
        
            if hasta:
                query += " AND DATE(fecha) <= ?"
                params.append(hasta)
        
            if buscar:
                query += " AND (proveedor LIKE ? OR descripcion LIKE ?)"
                params.append(f'%{buscar}%')
                params.append(f'%{buscar}%')
        
            query += " ORDER BY fecha DESC LIMIT ?"
            params.append(limite)
        
            cursor.execute(query, params)
            compras = []
        
            for row in cursor.fetchall():
                compras.append({
                    'id': row['id'],
                    'tipo': row['tipo'],
                    'descripcion': row['descripcion'],
                    'monto': row['monto'],
                    'proveedor': row['proveedor'],
                    'metodo_pago': row['metodo_pago'],
                    'fecha': row['fecha']
                })
        
            conn.close()
            return compras
        except Exception as e:
            print(f"Error al obtener compras: {e}")
            return []

    def obtener_compra(self, compra_id):
        """Obtener una compra específica por ID"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            cursor.execute("SELECT * FROM compras WHERE id = ?", (compra_id,))
            row = cursor.fetchone()
        
            if row:
                compra = {
                    'id': row['id'],
                    'tipo': row['tipo'],
                    'descripcion': row['descripcion'],
                    'monto': row['monto'],
                    'proveedor': row['proveedor'],
                    'metodo_pago': row['metodo_pago'],
                    'fecha': row['fecha']
                }
                conn.close()
                return compra
        
            conn.close()
            return None
        except Exception as e:
            print(f"Error al obtener compra: {e}")
            return None

    def obtener_detalle_compra(self, compra_id):
        """Obtener detalles de una compra de productos"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            cursor.execute('''
                SELECT * FROM detalle_compras 
                WHERE compra_id = ?
                ORDER BY id
            ''', (compra_id,))
        
            detalles = []
            for row in cursor.fetchall():
                detalles.append({
                    'id': row['id'],
                    'producto_id': row['producto_id'],
                    'producto_nombre': row['producto_nombre'],
                    'cantidad': row['cantidad'],
                    'precio_unitario': row['precio_unitario'],
                    'subtotal': row['subtotal']
                })
        
            conn.close()
            return detalles
        except Exception as e:
            print(f"Error al obtener detalle de compra: {e}")
            return []

    def actualizar_compra(self, id, tipo, monto, proveedor, metodo_pago, descripcion=None, observaciones=None):
        """Actualizar una compra existente"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            cursor.execute('''
                UPDATE compras 
                SET tipo = ?, descripcion = ?, monto = ?, proveedor = ?, metodo_pago = ?
                WHERE id = ?
            ''', (tipo, descripcion, monto, proveedor, metodo_pago, id))
        
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error al actualizar compra: {e}")
            return False

    def eliminar_compra(self, compra_id):
        """Eliminar una compra y sus detalles"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            # Primero eliminar los detalles si existen
            cursor.execute("DELETE FROM detalle_compras WHERE compra_id = ?", (compra_id,))
        
            # Luego eliminar la compra
            cursor.execute("DELETE FROM compras WHERE id = ?", (compra_id,))
        
            conn.commit()
            conn.close()
            return True
        except Exception as e:
            print(f"Error al eliminar compra: {e}")
            return False

    def obtener_estadisticas_compras(self, desde=None, hasta=None):
        """Obtener estadísticas de compras"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            # Total hoy
            cursor.execute('''
                SELECT COALESCE(SUM(monto), 0) as total 
                FROM compras 
                WHERE DATE(fecha) = DATE('now', 'localtime')
            ''')
            total_hoy = cursor.fetchone()['total']
        
            # Total mes actual
            cursor.execute('''
                SELECT COALESCE(SUM(monto), 0) as total 
                FROM compras 
                WHERE strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime')
            ''')
            total_mes = cursor.fetchone()['total']
        
            # Total gastos (tipo = 'gastos')
            query_gastos = "SELECT COALESCE(SUM(monto), 0) as total FROM compras WHERE tipo = 'gastos'"
            params_gastos = []
        
            if desde:
                query_gastos += " AND DATE(fecha) >= ?"
                params_gastos.append(desde)
        
            if hasta:
                query_gastos += " AND DATE(fecha) <= ?"
                params_gastos.append(hasta)
        
            cursor.execute(query_gastos, params_gastos)
            total_gastos = cursor.fetchone()['total']
        
            # Total general con filtros
            query_general = "SELECT COALESCE(SUM(monto), 0) as total FROM compras WHERE 1=1"
            params_general = []
        
            if desde:
                query_general += " AND DATE(fecha) >= ?"
                params_general.append(desde)
        
            if hasta:
                query_general += " AND DATE(fecha) <= ?"
                params_general.append(hasta)
        
            cursor.execute(query_general, params_general)
            total_general = cursor.fetchone()['total']
        
            conn.close()
        
            return {
                'totalHoy': total_hoy,
                'totalMes': total_mes,
                'totalGastos': total_gastos,
                'totalGeneral': total_general
            }
        except Exception as e:
            print(f"Error al obtener estadísticas de compras: {e}")
            return {
                'totalHoy': 0,
                'totalMes': 0,
                'totalGastos': 0,
                'totalGeneral': 0
            }

    def exportar_compras_excel(self, compras, fecha_desde=None, fecha_hasta=None):
        """Exportar compras a archivo Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
        
            wb = Workbook()
            ws = wb.active
            ws.title = "Compras y Gastos"
        
            # Título
            ws['A1'] = 'REPORTE DE COMPRAS Y GASTOS'
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'] = f'Período: {fecha_desde or "Inicio"} - {fecha_hasta or "Actual"}'
            ws['A2'].font = Font(italic=True)
        
            # Encabezados
            headers = ['ID', 'Fecha', 'Tipo', 'Descripción', 'Proveedor', 'Método Pago', 'Monto']
            ws.append([])  # Línea vacía
            ws.append(headers)
        
            # Estilo para encabezados
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
        
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
        
            # Datos
            total = 0
            for compra in compras:
                ws.append([
                    compra['id'],
                    compra['fecha'],
                    compra['tipo'].capitalize(),
                    compra['descripcion'] or '-',
                    compra['proveedor'] or '-',
                    compra['metodo_pago'].capitalize(),
                    f"Bs. {compra['monto']:.2f}"
                ])
                total += compra['monto']
        
            # Total
            ws.append([])
            ws.append(['', '', '', '', '', 'TOTAL:', f'Bs. {total:.2f}'])
        
            # Última fila en negrita
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
        
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
        
            # Guardar archivo
            os.makedirs('exports', exist_ok=True)
            filename = f'exports/compras_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(filename)
        
            return filename
        except Exception as e:
            print(f"Error al exportar compras: {e}")
            return None


# ============================================
# CÓDIGO PARA AGREGAR A database.py
# ============================================
# INSTRUCCIONES:
# 1. Abre tu archivo utils/database.py
# 2. Busca donde terminen las funciones de compras (después de def exportar_compras_excel)
# 3. Pega todo este código DESPUÉS de esas funciones (antes del cierre de la clase)
# IMPORTANTE: Asegúrate de que la indentación esté correcta (todos los def deben estar dentro de la clase Database)
# ============================================

    # ===== FUNCIONES DE CRÉDITOS =====

    def crear_credito(self, venta_id, cliente_nombre, cliente_telefono, monto_total):
        """Crear un nuevo crédito"""
        conn = self.get_connection()
        cursor = conn.cursor()
    
        cursor.execute('''
            INSERT INTO creditos 
            (venta_id, cliente_nombre, cliente_telefono, monto_total, saldo_pendiente)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            venta_id,
            cliente_nombre,
            cliente_telefono,
            monto_total,
            monto_total
        ))
    
        credito_id = cursor.lastrowid
        conn.commit()
        conn.close()
    
        return credito_id
    
    
    def obtener_creditos(self, estado='todos', busqueda=''):
        """Obtener créditos con filtros"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        query = 'SELECT * FROM creditos WHERE 1=1'
        params = []
        
        if estado != 'todos':
            query += ' AND estado = ?'
            params.append(estado)
        
        if busqueda:
            query += ' AND (cliente_nombre LIKE ? OR cliente_telefono LIKE ?)'
            params.extend([f'%{busqueda}%', f'%{busqueda}%'])
        
        query += ' ORDER BY fecha_credito DESC'
        
        cursor.execute(query, params)
        creditos = cursor.fetchall()
        
        conn.close()
        return creditos
    
    def obtener_credito(self, id):
        """Obtener un crédito por ID"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM creditos WHERE id = ?', (id,))
        credito = cursor.fetchone()
        
        conn.close()
        return credito
    
    def obtener_pagos_credito(self, credito_id):
        """Obtener pagos de un crédito"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM pagos_creditos 
            WHERE credito_id = ? 
            ORDER BY fecha DESC
        ''', (credito_id,))
        pagos = cursor.fetchall()
        
        conn.close()
        return pagos
    
    def registrar_pago_credito(self, credito_id, data):
        """Registrar un pago de crédito"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        monto = float(data.get('monto'))
        
        # Insertar pago
        cursor.execute('''
            INSERT INTO pagos_creditos 
            (credito_id, monto, metodo_pago)
            VALUES (?, ?, ?)
        ''', (credito_id, monto, data.get('metodo_pago', 'efectivo')))
        
        pago_id = cursor.lastrowid
        
        # Actualizar crédito
        cursor.execute('''
            UPDATE creditos 
            SET monto_pagado = monto_pagado + ?,
                saldo_pendiente = saldo_pendiente - ?,
                fecha_ultimo_pago = CURRENT_TIMESTAMP
            WHERE id = ?
        ''', (monto, monto, credito_id))
        
        # Verificar si está pagado completamente
        cursor.execute('SELECT saldo_pendiente FROM creditos WHERE id = ?', (credito_id,))
        saldo = cursor.fetchone()['saldo_pendiente']
        
        if saldo <= 0:
            cursor.execute('''
                UPDATE creditos 
                SET estado = 'pagado', saldo_pendiente = 0 
                WHERE id = ?
            ''', (credito_id,))
        elif saldo < cursor.execute('SELECT monto_total FROM creditos WHERE id = ?', (credito_id,)).fetchone()['monto_total']:
            cursor.execute('''
                UPDATE creditos 
                SET estado = 'parcial' 
                WHERE id = ?
            ''', (credito_id,))
        
        # Registrar movimiento de caja si hay caja abierta
        cursor.execute('SELECT id FROM caja WHERE estado = "abierta"')
        caja_abierta = cursor.fetchone()
        
        if caja_abierta:
            cursor.execute('''
                INSERT INTO movimientos_caja 
                (caja_id, tipo, concepto, monto, metodo_pago, referencia_id, referencia_tipo)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                caja_abierta['id'],
                'ingreso',
                f"Pago de crédito #{credito_id}",
                monto,
                data.get('metodo_pago', 'efectivo'),
                pago_id,
                'pago_credito'
            ))
        
        conn.commit()
        conn.close()
        
        return pago_id
    
    def obtener_estadisticas_creditos(self):
        """Obtener estadísticas de créditos"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Total créditos pendientes
        cursor.execute('''
            SELECT COUNT(*) as cantidad, COALESCE(SUM(saldo_pendiente), 0) as total
            FROM creditos 
            WHERE estado IN ('pendiente', 'parcial')
        ''')
        pendientes = cursor.fetchone()
        
        # Créditos vencidos (más de 30 días)
        cursor.execute('''
            SELECT COUNT(*) as cantidad, COALESCE(SUM(saldo_pendiente), 0) as total
            FROM creditos 
            WHERE estado IN ('pendiente', 'parcial')
            AND date(fecha_credito) <= date('now', '-30 days')
        ''')
        vencidos = cursor.fetchone()
        
        # Total cobrado este mes
        cursor.execute('''
            SELECT COALESCE(SUM(monto), 0) as total
            FROM pagos_creditos
            WHERE strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now', 'localtime')
        ''')
        cobrado_mes = cursor.fetchone()
        
        # Total clientes con crédito
        cursor.execute('''
            SELECT COUNT(DISTINCT cliente_nombre) as total
            FROM creditos
            WHERE estado IN ('pendiente', 'parcial')
        ''')
        clientes = cursor.fetchone()
        
        conn.close()
        
        return {
            'creditos_pendientes': pendientes['cantidad'],
            'total_pendiente': pendientes['total'],
            'creditos_vencidos': vencidos['cantidad'],
            'total_vencido': vencidos['total'],
            'cobrado_mes': cobrado_mes['total'],
            'total_clientes': clientes['total']
        }
    
    def obtener_resumen_creditos_cliente(self, nombre_cliente):
        """Obtener resumen de créditos de un cliente"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Créditos del cliente
        cursor.execute('''
            SELECT * FROM creditos 
            WHERE cliente_nombre = ? 
            ORDER BY fecha_credito DESC
        ''', (nombre_cliente,))
        creditos = cursor.fetchall()
        
        # Total adeudado
        cursor.execute('''
            SELECT COALESCE(SUM(saldo_pendiente), 0) as total
            FROM creditos 
            WHERE cliente_nombre = ? 
            AND estado IN ('pendiente', 'parcial')
        ''', (nombre_cliente,))
        total_adeudado = cursor.fetchone()['total']
        
        conn.close()
        
        return {
            'creditos': [dict(c) for c in creditos],
            'total_adeudado': total_adeudado
        }
    
    def exportar_creditos_excel(self):
        """Exportar créditos a Excel"""
        creditos = self.obtener_creditos()
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Créditos"
        
        # Encabezados
        headers = ['ID', 'Fecha', 'Cliente', 'Teléfono', 'Monto Total', 
                   'Monto Pagado', 'Saldo Pendiente', 'Estado', 'Último Pago']
        ws.append(headers)
        
        # Estilos
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
            cell.font = Font(color="000000", bold=True)
        
        # Datos
        for credito in creditos:
            ws.append([
                credito['id'],
                credito['fecha_credito'],
                credito['cliente_nombre'],
                credito['cliente_telefono'] or '-',
                credito['monto_total'],
                credito['monto_pagado'],
                credito['saldo_pendiente'],
                credito['estado'],
                credito['fecha_ultimo_pago'] or '-'
            ])
        
        # Ajustar columnas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Guardar
        filename = f"exports/creditos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename
    
    # ===== FUNCIONES DE CAJA =====

    # ============================================
# CORRECCIÓN PARA database.py - abrir_caja
# ============================================
# BUSCA la función abrir_caja y REEMPLÁZALA con este código:
# ============================================

    def abrir_caja(self, monto_inicial=0):
        """Abrir una nueva caja"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
        
            # Verificar que no haya caja abierta
            cursor.execute('SELECT id FROM caja WHERE estado = "abierta"')
            if cursor.fetchone():
                conn.close()
                raise Exception('Ya hay una caja abierta')
        
            # Crear nueva caja
            cursor.execute('''
                INSERT INTO caja (monto_inicial, usuario)
                VALUES (?, ?)
            ''', (float(monto_inicial), 'admin'))
        
            caja_id = cursor.lastrowid
        
            # NO registrar movimiento inicial (el monto_inicial ya está en la tabla caja)
            # Esto evita la duplicación
        
            conn.commit()
            conn.close()
        
            print(f'✅ Caja abierta: ID {caja_id} - Monto inicial: Bs. {monto_inicial}')
            return caja_id
        except Exception as e:
            print(f'❌ Error al abrir caja: {e}')
            raise e
    
    
    
    def cerrar_caja(self, caja_id, efectivo_contado):
        """Cerrar caja actual"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Calcular totales
        cursor.execute('''
            SELECT 
                COALESCE(SUM(CASE WHEN tipo = 'ingreso' AND metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as total_efectivo_ing,
                COALESCE(SUM(CASE WHEN tipo = 'egreso' AND metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as total_efectivo_egr,
                COALESCE(SUM(CASE WHEN tipo = 'ingreso' AND metodo_pago = 'qr' THEN monto ELSE 0 END), 0) as total_qr,
                COALESCE(SUM(CASE WHEN tipo = 'ingreso' AND metodo_pago = 'credito' THEN monto ELSE 0 END), 0) as total_credito,
                COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                COALESCE(SUM(CASE WHEN tipo = 'egreso' THEN monto ELSE 0 END), 0) as total_egresos
            FROM movimientos_caja 
            WHERE caja_id = ?
        ''', (caja_id,))
        
        totales = cursor.fetchone()
        
        # Obtener monto inicial
        cursor.execute('SELECT monto_inicial FROM caja WHERE id = ?', (caja_id,))
        monto_inicial = cursor.fetchone()['monto_inicial']
        
        # Calcular efectivo esperado
        efectivo_esperado = monto_inicial + totales['total_efectivo_ing'] - totales['total_efectivo_egr']
        diferencia = float(efectivo_contado) - efectivo_esperado
        
        # Actualizar caja
        cursor.execute('''
            UPDATE caja SET
                fecha_cierre = CURRENT_TIMESTAMP,
                total_efectivo = ?,
                total_qr = ?,
                total_credito = ?,
                total_ingresos = ?,
                total_egresos = ?,
                efectivo_esperado = ?,
                efectivo_contado = ?,
                diferencia = ?,
                estado = 'cerrada'
            WHERE id = ?
        ''', (
            totales['total_efectivo_ing'] - totales['total_efectivo_egr'],
            totales['total_qr'],
            totales['total_credito'],
            totales['total_ingresos'],
            totales['total_egresos'],
            efectivo_esperado,
            efectivo_contado,
            diferencia,
            caja_id
        ))
        
        conn.commit()
        conn.close()
        
        return {
            'efectivo_esperado': efectivo_esperado,
            'efectivo_contado': efectivo_contado,
            'diferencia': diferencia,
            'total_ingresos': totales['total_ingresos'],
            'total_egresos': totales['total_egresos']
        }
    
    def obtener_caja_actual(self):
        """Obtener la caja abierta actual"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM caja WHERE estado = "abierta"')
        caja = cursor.fetchone()
        
        conn.close()
        return caja
    
    def obtener_caja(self, id):
        """Obtener una caja por ID"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM caja WHERE id = ?', (id,))
        caja = cursor.fetchone()
        
        conn.close()
        return caja
    
    def obtener_movimientos_caja(self, caja_id):
        """Obtener movimientos de una caja"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT * FROM movimientos_caja 
            WHERE caja_id = ? 
            ORDER BY fecha DESC
        ''', (caja_id,))
        movimientos = cursor.fetchall()
        
        conn.close()
        return movimientos
    
    ## 3️⃣ FUNCIÓN obtener_resumen_caja - CORREGIDA (para calcular bien el efectivo):

# ============================================
# CORRECCIÓN URGENTE PARA database.py
# ============================================
# BUSCA la función obtener_resumen_caja (alrededor de línea 800-900)
# Y REEMPLÁZALA COMPLETA con este código:
# ============================================

    def obtener_resumen_caja(self, caja_id):
        """Obtener resumen de una caja"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Obtener caja
            cursor.execute('SELECT * FROM caja WHERE id = ?', (caja_id,))
            caja = cursor.fetchone()
            
            if not caja:
                conn.close()
                return {}
            
            # Calcular totales por tipo y método
            cursor.execute('''
                SELECT 
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' AND metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as ingresos_efectivo,
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' AND metodo_pago = 'qr' THEN monto ELSE 0 END), 0) as ingresos_qr,
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' AND metodo_pago = 'mixto' THEN monto ELSE 0 END), 0) as ingresos_mixto,
                    COALESCE(SUM(CASE WHEN tipo = 'egreso' AND metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as egresos_efectivo,
                    COALESCE(SUM(CASE WHEN tipo = 'egreso' AND metodo_pago != 'efectivo' THEN monto ELSE 0 END), 0) as egresos_otros,
                    COALESCE(SUM(CASE WHEN tipo = 'ingreso' THEN monto ELSE 0 END), 0) as total_ingresos,
                    COALESCE(SUM(CASE WHEN tipo = 'egreso' THEN monto ELSE 0 END), 0) as total_egresos,
                    COUNT(CASE WHEN referencia_tipo = 'venta' THEN 1 END) as num_ventas,
                    COUNT(CASE WHEN referencia_tipo = 'compra' THEN 1 END) as num_compras,
                    COUNT(CASE WHEN referencia_tipo = 'pago_credito' THEN 1 END) as num_pagos
                FROM movimientos_caja 
                WHERE caja_id = ?
            ''', (caja_id,))
            
            totales = cursor.fetchone()
            
            # Calcular efectivo actual (monto inicial + ingresos efectivo - egresos efectivo)
            efectivo_actual = caja['monto_inicial'] + totales['ingresos_efectivo'] - totales['egresos_efectivo']
            
            conn.close()
            
            return {
                'monto_inicial': caja['monto_inicial'],
                'efectivo_actual': efectivo_actual,
                'ingresos_efectivo': totales['ingresos_efectivo'],
                'ingresos_qr': totales['ingresos_qr'],
                'ingresos_mixto': totales['ingresos_mixto'],
                'egresos_efectivo': totales['egresos_efectivo'],
                'egresos_otros': totales['egresos_otros'],
                'total_ingresos': totales['total_ingresos'],
                'total_egresos': totales['total_egresos'],
                'num_ventas': totales['num_ventas'],
                'num_compras': totales['num_compras'],
                'num_pagos': totales['num_pagos'],
                'balance': totales['total_ingresos'] - totales['total_egresos']
            }
        except Exception as e:
            print(f'❌ Error en obtener_resumen_caja: {e}')
            return {
                'monto_inicial': 0,
                'efectivo_actual': 0,
                'ingresos_efectivo': 0,
                'ingresos_qr': 0,
                'ingresos_mixto': 0,
                'egresos_efectivo': 0,
                'egresos_otros': 0,
                'total_ingresos': 0,
                'total_egresos': 0,
                'num_ventas': 0,
                'num_compras': 0,
                'num_pagos': 0,
                'balance': 0
            }

# ====================================
# INSTRUCCIONES:
# 1. Abre database.py
# 2. Busca y REEMPLAZA estas 3 funciones completas:
#    - abrir_caja
#    - crear_venta  
#    - obtener_resumen_caja
# 3. Guarda el archivo
# 4. Reinicia el servidor (Ctrl+C y python app.py)
# ====================================
    
    
    
    def obtener_historial_cajas(self, fecha_desde=None, fecha_hasta=None):
        """Obtener historial de cajas"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        query = 'SELECT * FROM caja WHERE estado = "cerrada"'
        params = []
        
        if fecha_desde:
            query += ' AND date(fecha_apertura) >= date(?)'
            params.append(fecha_desde)
        
        if fecha_hasta:
            query += ' AND date(fecha_apertura) <= date(?)'
            params.append(fecha_hasta)
        
        query += ' ORDER BY fecha_apertura DESC'
        
        cursor.execute(query, params)
        cajas = cursor.fetchall()
        
        conn.close()
        return cajas
    
    def registrar_retiro_caja(self, caja_id, data):
        """Registrar un retiro de caja"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO movimientos_caja 
            (caja_id, tipo, concepto, monto, metodo_pago)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            caja_id,
            'egreso',
            data.get('concepto', 'Retiro de efectivo'),
            float(data.get('monto')),
            'efectivo'
        ))
        
        movimiento_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return movimiento_id
    
    def exportar_caja_excel(self, caja_id):
        """Exportar reporte de caja a Excel"""
        caja = self.obtener_caja(caja_id)
        movimientos = self.obtener_movimientos_caja(caja_id)
        resumen = self.obtener_resumen_caja(caja_id)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        
        # Hoja 1: Resumen
        ws1 = wb.active
        ws1.title = "Resumen"
        
        ws1.append(['REPORTE DE CAJA'])
        ws1.append([''])
        ws1.append(['Fecha Apertura:', caja['fecha_apertura']])
        ws1.append(['Fecha Cierre:', caja['fecha_cierre'] or 'Abierta'])
        ws1.append(['Usuario:', caja['usuario']])
        ws1.append([''])
        ws1.append(['RESUMEN FINANCIERO'])
        ws1.append(['Monto Inicial:', f"Bs. {resumen['monto_inicial']:.2f}"])
        ws1.append(['Total Ingresos:', f"Bs. {resumen['total_ingresos']:.2f}"])
        ws1.append(['Total Egresos:', f"Bs. {resumen['total_egresos']:.2f}"])
        ws1.append(['Balance:', f"Bs. {resumen['balance']:.2f}"])
        ws1.append([''])
        
        if caja['estado'] == 'cerrada':
            ws1.append(['CIERRE DE CAJA'])
            ws1.append(['Efectivo Esperado:', f"Bs. {caja['efectivo_esperado']:.2f}"])
            ws1.append(['Efectivo Contado:', f"Bs. {caja['efectivo_contado']:.2f}"])
            ws1.append(['Diferencia:', f"Bs. {caja['diferencia']:.2f}"])
        
        # Hoja 2: Movimientos
        ws2 = wb.create_sheet("Movimientos")
        
        headers = ['Fecha', 'Tipo', 'Concepto', 'Monto', 'Método Pago']
        ws2.append(headers)
        
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for mov in movimientos:
            ws2.append([
                mov['fecha'],
                mov['tipo'],
                mov['concepto'],
                mov['monto'],
                mov['metodo_pago'] or '-'
            ])
        
        # Ajustar columnas
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Guardar
        filename = f"exports/caja_{caja_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename
    
    def exportar_historial_cajas_excel(self, fecha_desde=None, fecha_hasta=None):
        """Exportar historial de cajas a Excel"""
        cajas = self.obtener_historial_cajas(fecha_desde, fecha_hasta)
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Historial Cajas"
        
        # Encabezados
        headers = ['ID', 'Apertura', 'Cierre', 'Inicial', 'Ingresos', 
                   'Egresos', 'Esperado', 'Contado', 'Diferencia']
        ws.append(headers)
        
        # Estilos
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="6c757d", end_color="6c757d", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Datos
        for caja in cajas:
            ws.append([
                caja['id'],
                caja['fecha_apertura'],
                caja['fecha_cierre'],
                caja['monto_inicial'],
                caja['total_ingresos'],
                caja['total_egresos'],
                caja['efectivo_esperado'],
                caja['efectivo_contado'],
                caja['diferencia']
            ])
        
        # Ajustar columnas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Guardar
        filename = f"exports/historial_cajas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)
        
        return filename

# ============================================
# FIN DEL CÓDIGO PARA database.py
# ============================================
# IMPORTANTE: Verifica que la indentación sea correcta
# Todos los 'def' deben estar al mismo nivel que las otras funciones dentro de la clase Database
# ============================================

# ============================================
# FUNCIONES DE ESTADÍSTICAS PARA database.py
# ============================================
# INSTRUCCIONES:
# 1. Abre tu archivo database.py
# 2. Busca el final del archivo (antes de "if __name__ == '__main__':")
# 3. Copia y pega TODAS estas funciones dentro de la clase Database
# 4. Asegúrate de que la indentación sea correcta (4 espacios)
# ============================================

    # ========== FUNCIONES PARA ESTADÍSTICAS (MÓDULO 7) ==========
    
    def obtener_estadisticas_dashboard(self):
        """Obtener estadísticas para el dashboard principal"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            hoy = datetime.now().strftime('%Y-%m-%d')
            
            # Total productos
            cursor.execute('SELECT COUNT(*) as total FROM productos')
            total_productos = cursor.fetchone()['total']
            
            # Productos con stock bajo
            cursor.execute('SELECT COUNT(*) as total FROM productos WHERE stock <= stock_minimo')
            stock_bajo = cursor.fetchone()['total']
            
            # Ventas de hoy
            cursor.execute('''
                SELECT COUNT(*) as cantidad, COALESCE(SUM(total), 0) as total 
                FROM ventas 
                WHERE DATE(fecha) = DATE(?)
            ''', (hoy,))
            ventas_hoy = cursor.fetchone()
            
            # Ventas del mes
            cursor.execute('''
                SELECT COUNT(*) as cantidad, COALESCE(SUM(total), 0) as total 
                FROM ventas 
                WHERE strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')
            ''')
            ventas_mes = cursor.fetchone()
            
            # Compras/Gastos del mes
            cursor.execute('''
                SELECT COALESCE(SUM(monto), 0) as total 
                FROM compras 
                WHERE strftime('%Y-%m', fecha) = strftime('%Y-%m', 'now')
            ''')
            gastos_mes = cursor.fetchone()['total']
            
            # Créditos pendientes
            cursor.execute('''
                SELECT COUNT(*) as cantidad, COALESCE(SUM(saldo_pendiente), 0) as total 
                FROM creditos 
                WHERE estado IN ('pendiente', 'parcial')
            ''')
            creditos = cursor.fetchone()
            
            # Estado de caja
            cursor.execute('SELECT * FROM caja WHERE estado = "abierta"')
            caja_abierta = cursor.fetchone()
            
            caja_estado = {
                'abierta': caja_abierta is not None,
                'monto_actual': 0
            }
            
            if caja_abierta:
                # Calcular efectivo actual
                cursor.execute('''
                    SELECT 
                        COALESCE(SUM(CASE WHEN tipo = 'ingreso' AND metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) -
                        COALESCE(SUM(CASE WHEN tipo = 'egreso' AND metodo_pago = 'efectivo' THEN monto ELSE 0 END), 0) as efectivo
                    FROM movimientos_caja
                    WHERE caja_id = ?
                ''', (caja_abierta['id'],))
                movimientos = cursor.fetchone()
                caja_estado['monto_actual'] = caja_abierta['monto_inicial'] + (movimientos['efectivo'] if movimientos else 0)
            
            conn.close()
            
            return {
                'productos': {
                    'total': total_productos,
                    'stock_bajo': stock_bajo
                },
                'ventas_hoy': {
                    'cantidad': ventas_hoy['cantidad'],
                    'total': ventas_hoy['total']
                },
                'ventas_mes': {
                    'cantidad': ventas_mes['cantidad'],
                    'total': ventas_mes['total']
                },
                'gastos_mes': gastos_mes,
                'creditos': {
                    'cantidad': creditos['cantidad'],
                    'total': creditos['total']
                },
                'caja': caja_estado,
                'ganancia_mes': ventas_mes['total'] - gastos_mes
            }
            
        except Exception as e:
            print(f'❌ Error al obtener estadísticas dashboard: {e}')
            return None
    
    def obtener_ventas_por_periodo(self, periodo='dia', fecha_desde=None, fecha_hasta=None):
        """Obtener ventas agrupadas por período (día, semana, mes)"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Si no hay fechas, usar últimos 30 días
            if not fecha_desde:
                fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            if not fecha_hasta:
                fecha_hasta = datetime.now().strftime('%Y-%m-%d')
            
            if periodo == 'dia':
                cursor.execute('''
                    SELECT 
                        DATE(fecha) as periodo,
                        COUNT(*) as cantidad,
                        COALESCE(SUM(total), 0) as total,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'efectivo' THEN total ELSE 0 END), 0) as efectivo,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'qr' THEN total ELSE 0 END), 0) as qr,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'credito' THEN total ELSE 0 END), 0) as credito,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'mixto' THEN total ELSE 0 END), 0) as mixto
                    FROM ventas
                    WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
                    GROUP BY DATE(fecha)
                    ORDER BY DATE(fecha) ASC
                ''', (fecha_desde, fecha_hasta))
            
            elif periodo == 'semana':
                cursor.execute('''
                    SELECT 
                        strftime('%Y-W%W', fecha) as periodo,
                        COUNT(*) as cantidad,
                        COALESCE(SUM(total), 0) as total,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'efectivo' THEN total ELSE 0 END), 0) as efectivo,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'qr' THEN total ELSE 0 END), 0) as qr,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'credito' THEN total ELSE 0 END), 0) as credito,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'mixto' THEN total ELSE 0 END), 0) as mixto
                    FROM ventas
                    WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
                    GROUP BY strftime('%Y-W%W', fecha)
                    ORDER BY periodo ASC
                ''', (fecha_desde, fecha_hasta))
            
            else:  # mes
                cursor.execute('''
                    SELECT 
                        strftime('%Y-%m', fecha) as periodo,
                        COUNT(*) as cantidad,
                        COALESCE(SUM(total), 0) as total,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'efectivo' THEN total ELSE 0 END), 0) as efectivo,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'qr' THEN total ELSE 0 END), 0) as qr,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'credito' THEN total ELSE 0 END), 0) as credito,
                        COALESCE(SUM(CASE WHEN metodo_pago = 'mixto' THEN total ELSE 0 END), 0) as mixto
                    FROM ventas
                    WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
                    GROUP BY strftime('%Y-%m', fecha)
                    ORDER BY periodo ASC
                ''', (fecha_desde, fecha_hasta))
            
            rows = cursor.fetchall()
            conn.close()
            
            return [dict(row) for row in rows]
            
        except Exception as e:
            print(f'❌ Error al obtener ventas por período: {e}')
            return []
    
    def obtener_compras_por_periodo(self, periodo='dia', fecha_desde=None, fecha_hasta=None):
        """Obtener compras/gastos agrupados por período"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            if not fecha_desde:
                fecha_desde = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
            if not fecha_hasta:
                fecha_hasta = datetime.now().strftime('%Y-%m-%d')
            
            if periodo == 'dia':
                cursor.execute('''
                    SELECT 
                        DATE(fecha) as periodo,
                        COALESCE(SUM(CASE WHEN tipo = 'productos' THEN monto ELSE 0 END), 0) as productos,
                        COALESCE(SUM(CASE WHEN tipo = 'insumos' THEN monto ELSE 0 END), 0) as insumos,
                        COALESCE(SUM(CASE WHEN tipo = 'gastos' THEN monto ELSE 0 END), 0) as gastos,
                        COALESCE(SUM(monto), 0) as total
                    FROM compras
                    WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
                    GROUP BY DATE(fecha)
                    ORDER BY DATE(fecha) ASC
                ''', (fecha_desde, fecha_hasta))
            else:
                cursor.execute('''
                    SELECT 
                        strftime('%Y-%m', fecha) as periodo,
                        COALESCE(SUM(CASE WHEN tipo = 'productos' THEN monto ELSE 0 END), 0) as productos,
                        COALESCE(SUM(CASE WHEN tipo = 'insumos' THEN monto ELSE 0 END), 0) as insumos,
                        COALESCE(SUM(CASE WHEN tipo = 'gastos' THEN monto ELSE 0 END), 0) as gastos,
                        COALESCE(SUM(monto), 0) as total
                    FROM compras
                    WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
                    GROUP BY strftime('%Y-%m', fecha)
                    ORDER BY periodo ASC
                ''', (fecha_desde, fecha_hasta))
            
            rows = cursor.fetchall()
            conn.close()
            
            return [dict(row) for row in rows]
            
        except Exception as e:
            print(f'❌ Error al obtener compras por período: {e}')
            return []
    
    def obtener_top_productos(self, limite=10, fecha_desde=None, fecha_hasta=None):
        """Obtener los productos más vendidos"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            query = '''
                SELECT 
                    dv.producto_id,
                    dv.producto_nombre,
                    SUM(dv.cantidad) as cantidad_vendida,
                    SUM(dv.subtotal) as total_vendido,
                    p.categoria,
                    p.stock
                FROM detalle_ventas dv
                LEFT JOIN productos p ON dv.producto_id = p.id
                LEFT JOIN ventas v ON dv.venta_id = v.id
            '''
            
            params = []
            if fecha_desde and fecha_hasta:
                query += ' WHERE DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)'
                params = [fecha_desde, fecha_hasta]
            
            query += '''
                GROUP BY dv.producto_id, dv.producto_nombre
                ORDER BY cantidad_vendida DESC
                LIMIT ?
            '''
            params.append(limite)
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()
            
            return [dict(row) for row in rows]
            
        except Exception as e:
            print(f'❌ Error al obtener top productos: {e}')
            return []
    
    def obtener_ventas_por_categoria(self, fecha_desde=None, fecha_hasta=None):
        """Obtener ventas agrupadas por categoría de producto"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            query = '''
                SELECT 
                    p.categoria,
                    SUM(dv.cantidad) as cantidad,
                    SUM(dv.subtotal) as total
                FROM detalle_ventas dv
                LEFT JOIN productos p ON dv.producto_id = p.id
                LEFT JOIN ventas v ON dv.venta_id = v.id
            '''
            
            params = []
            if fecha_desde and fecha_hasta:
                query += ' WHERE DATE(v.fecha) BETWEEN DATE(?) AND DATE(?)'
                params = [fecha_desde, fecha_hasta]
            
            query += ' GROUP BY p.categoria ORDER BY total DESC'
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()
            
            return [dict(row) for row in rows]
            
        except Exception as e:
            print(f'❌ Error al obtener ventas por categoría: {e}')
            return []
    
    def obtener_resumen_financiero(self, fecha_desde=None, fecha_hasta=None):
        """Obtener resumen financiero completo"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            if not fecha_desde:
                fecha_desde = datetime.now().replace(day=1).strftime('%Y-%m-%d')
            if not fecha_hasta:
                fecha_hasta = datetime.now().strftime('%Y-%m-%d')
            
            # Total ventas
            cursor.execute('''
                SELECT 
                    COUNT(*) as cantidad,
                    COALESCE(SUM(total), 0) as total,
                    COALESCE(SUM(CASE WHEN metodo_pago = 'efectivo' THEN total ELSE 0 END), 0) as efectivo,
                    COALESCE(SUM(CASE WHEN metodo_pago = 'qr' THEN total ELSE 0 END), 0) as qr,
                    COALESCE(SUM(CASE WHEN metodo_pago = 'credito' THEN total ELSE 0 END), 0) as credito,
                    COALESCE(SUM(CASE WHEN metodo_pago = 'mixto' THEN monto_efectivo ELSE 0 END), 0) as mixto_efectivo,
                    COALESCE(SUM(CASE WHEN metodo_pago = 'mixto' THEN monto_qr ELSE 0 END), 0) as mixto_qr
                FROM ventas
                WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
            ''', (fecha_desde, fecha_hasta))
            ventas = cursor.fetchone()
            
            # Total compras por tipo
            cursor.execute('''
                SELECT 
                    COUNT(*) as cantidad,
                    COALESCE(SUM(monto), 0) as total,
                    COALESCE(SUM(CASE WHEN tipo = 'productos' THEN monto ELSE 0 END), 0) as productos,
                    COALESCE(SUM(CASE WHEN tipo = 'insumos' THEN monto ELSE 0 END), 0) as insumos,
                    COALESCE(SUM(CASE WHEN tipo = 'gastos' THEN monto ELSE 0 END), 0) as gastos
                FROM compras
                WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
            ''', (fecha_desde, fecha_hasta))
            compras = cursor.fetchone()
            
            # Créditos cobrados en el período
            cursor.execute('''
                SELECT COALESCE(SUM(monto), 0) as total
                FROM pagos_creditos
                WHERE DATE(fecha) BETWEEN DATE(?) AND DATE(?)
            ''', (fecha_desde, fecha_hasta))
            pagos_creditos = cursor.fetchone()['total']
            
            # Créditos otorgados en el período
            cursor.execute('''
                SELECT COALESCE(SUM(monto_total), 0) as total
                FROM creditos
                WHERE DATE(fecha_credito) BETWEEN DATE(?) AND DATE(?)
            ''', (fecha_desde, fecha_hasta))
            creditos_otorgados = cursor.fetchone()['total']
            
            conn.close()
            
            total_ingresos = ventas['total'] + pagos_creditos
            total_egresos = compras['total']
            ganancia_bruta = total_ingresos - total_egresos
            
            return {
                'periodo': {
                    'desde': fecha_desde,
                    'hasta': fecha_hasta
                },
                'ventas': {
                    'cantidad': ventas['cantidad'],
                    'total': ventas['total'],
                    'efectivo': ventas['efectivo'] + ventas['mixto_efectivo'],
                    'qr': ventas['qr'] + ventas['mixto_qr'],
                    'credito': ventas['credito']
                },
                'compras': {
                    'cantidad': compras['cantidad'],
                    'total': compras['total'],
                    'productos': compras['productos'],
                    'insumos': compras['insumos'],
                    'gastos': compras['gastos']
                },
                'creditos': {
                    'otorgados': creditos_otorgados,
                    'cobrados': pagos_creditos
                },
                'resumen': {
                    'ingresos': total_ingresos,
                    'egresos': total_egresos,
                    'ganancia_bruta': ganancia_bruta
                }
            }
            
        except Exception as e:
            print(f'❌ Error al obtener resumen financiero: {e}')
            return None
    
    def obtener_comparativa_periodos(self):
        """Comparar ventas entre períodos (hoy vs ayer, esta semana vs anterior, este mes vs anterior)"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            hoy = datetime.now()
            ayer = hoy - timedelta(days=1)
            inicio_semana = hoy - timedelta(days=hoy.weekday())
            inicio_semana_anterior = inicio_semana - timedelta(days=7)
            fin_semana_anterior = inicio_semana - timedelta(days=1)
            inicio_mes = hoy.replace(day=1)
            if hoy.month == 1:
                inicio_mes_anterior = hoy.replace(year=hoy.year-1, month=12, day=1)
            else:
                inicio_mes_anterior = hoy.replace(month=hoy.month-1, day=1)
            fin_mes_anterior = inicio_mes - timedelta(days=1)
            
            # Ventas de hoy
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as cantidad
                FROM ventas WHERE DATE(fecha) = DATE(?)
            ''', (hoy.strftime('%Y-%m-%d'),))
            ventas_hoy = cursor.fetchone()
            
            # Ventas de ayer
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as cantidad
                FROM ventas WHERE DATE(fecha) = DATE(?)
            ''', (ayer.strftime('%Y-%m-%d'),))
            ventas_ayer = cursor.fetchone()
            
            # Ventas esta semana
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as cantidad
                FROM ventas WHERE DATE(fecha) >= DATE(?) AND DATE(fecha) <= DATE(?)
            ''', (inicio_semana.strftime('%Y-%m-%d'), hoy.strftime('%Y-%m-%d')))
            ventas_semana = cursor.fetchone()
            
            # Ventas semana anterior
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as cantidad
                FROM ventas WHERE DATE(fecha) >= DATE(?) AND DATE(fecha) <= DATE(?)
            ''', (inicio_semana_anterior.strftime('%Y-%m-%d'), fin_semana_anterior.strftime('%Y-%m-%d')))
            ventas_semana_anterior = cursor.fetchone()
            
            # Ventas este mes
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as cantidad
                FROM ventas WHERE DATE(fecha) >= DATE(?) AND DATE(fecha) <= DATE(?)
            ''', (inicio_mes.strftime('%Y-%m-%d'), hoy.strftime('%Y-%m-%d')))
            ventas_mes = cursor.fetchone()
            
            # Ventas mes anterior
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total, COUNT(*) as cantidad
                FROM ventas WHERE DATE(fecha) >= DATE(?) AND DATE(fecha) <= DATE(?)
            ''', (inicio_mes_anterior.strftime('%Y-%m-%d'), fin_mes_anterior.strftime('%Y-%m-%d')))
            ventas_mes_anterior = cursor.fetchone()
            
            conn.close()
            
            # Calcular variaciones
            def calcular_variacion(actual, anterior):
                if anterior == 0:
                    return 100 if actual > 0 else 0
                return round(((actual - anterior) / anterior) * 100, 1)
            
            return {
                'hoy': {
                    'total': ventas_hoy['total'],
                    'cantidad': ventas_hoy['cantidad'],
                    'comparar_con': ventas_ayer['total'],
                    'variacion': calcular_variacion(ventas_hoy['total'], ventas_ayer['total'])
                },
                'semana': {
                    'total': ventas_semana['total'],
                    'cantidad': ventas_semana['cantidad'],
                    'comparar_con': ventas_semana_anterior['total'],
                    'variacion': calcular_variacion(ventas_semana['total'], ventas_semana_anterior['total'])
                },
                'mes': {
                    'total': ventas_mes['total'],
                    'cantidad': ventas_mes['cantidad'],
                    'comparar_con': ventas_mes_anterior['total'],
                    'variacion': calcular_variacion(ventas_mes['total'], ventas_mes_anterior['total'])
                }
            }
            
        except Exception as e:
            print(f'❌ Error al obtener comparativa: {e}')
            return None
    
    def obtener_ventas_por_hora(self, fecha=None):
        """Obtener distribución de ventas por hora del día"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            if not fecha:
                fecha = datetime.now().strftime('%Y-%m-%d')
            
            cursor.execute('''
                SELECT 
                    CAST(strftime('%H', fecha) AS INTEGER) as hora,
                    COUNT(*) as cantidad,
                    COALESCE(SUM(total), 0) as total
                FROM ventas
                WHERE DATE(fecha) = DATE(?)
                GROUP BY CAST(strftime('%H', fecha) AS INTEGER)
                ORDER BY hora ASC
            ''', (fecha,))
            
            rows = cursor.fetchall()
            conn.close()
            
            # Completar todas las horas del día (6 AM - 11 PM)
            resultado = {}
            for hora in range(6, 24):
                resultado[hora] = {'cantidad': 0, 'total': 0}
            
            for row in rows:
                if row['hora'] in resultado:
                    resultado[row['hora']] = {
                        'cantidad': row['cantidad'],
                        'total': row['total']
                    }
            
            return resultado
            
        except Exception as e:
            print(f'❌ Error al obtener ventas por hora: {e}')
            return {}
    
    def exportar_estadisticas_excel(self, fecha_desde, fecha_hasta):
        """Exportar reporte de estadísticas completo a Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import BarChart, PieChart, LineChart, Reference
            
            wb = openpyxl.Workbook()
            
            # ===== HOJA 1: RESUMEN EJECUTIVO =====
            ws1 = wb.active
            ws1.title = "Resumen Ejecutivo"
            
            resumen = self.obtener_resumen_financiero(fecha_desde, fecha_hasta)
            
            # Título
            ws1['A1'] = 'REPORTE DE ESTADÍSTICAS - BEER LICORERÍA'
            ws1['A1'].font = Font(size=16, bold=True, color="FFFFFF")
            ws1['A1'].fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
            ws1.merge_cells('A1:E1')
            
            ws1['A2'] = f'Período: {fecha_desde} al {fecha_hasta}'
            ws1['A2'].font = Font(italic=True)
            ws1.merge_cells('A2:E2')
            
            # Resumen financiero
            ws1['A4'] = 'RESUMEN FINANCIERO'
            ws1['A4'].font = Font(bold=True, size=12)
            
            ws1['A5'] = 'Total Ventas:'
            ws1['B5'] = f"Bs. {resumen['ventas']['total']:.2f}" if resumen else "Bs. 0.00"
            ws1['A6'] = 'Total Compras/Gastos:'
            ws1['B6'] = f"Bs. {resumen['compras']['total']:.2f}" if resumen else "Bs. 0.00"
            ws1['A7'] = 'Ganancia Bruta:'
            ws1['B7'] = f"Bs. {resumen['resumen']['ganancia_bruta']:.2f}" if resumen else "Bs. 0.00"
            ws1['B7'].font = Font(bold=True, color="28a745" if resumen and resumen['resumen']['ganancia_bruta'] >= 0 else "dc3545")
            
            # ===== HOJA 2: TOP PRODUCTOS =====
            ws2 = wb.create_sheet("Top Productos")
            
            top_productos = self.obtener_top_productos(10, fecha_desde, fecha_hasta)
            
            headers = ['#', 'Producto', 'Categoría', 'Cantidad Vendida', 'Total Vendido']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")
            
            for idx, prod in enumerate(top_productos, 1):
                ws2.cell(row=idx+1, column=1, value=idx)
                ws2.cell(row=idx+1, column=2, value=prod['producto_nombre'])
                ws2.cell(row=idx+1, column=3, value=prod.get('categoria', '-'))
                ws2.cell(row=idx+1, column=4, value=prod['cantidad_vendida'])
                ws2.cell(row=idx+1, column=5, value=f"Bs. {prod['total_vendido']:.2f}")
            
            # ===== HOJA 3: VENTAS POR DÍA =====
            ws3 = wb.create_sheet("Ventas por Día")
            
            ventas_dia = self.obtener_ventas_por_periodo('dia', fecha_desde, fecha_hasta)
            
            headers = ['Fecha', 'Cantidad', 'Efectivo', 'QR', 'Crédito', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws3.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
            
            for idx, venta in enumerate(ventas_dia, 1):
                ws3.cell(row=idx+1, column=1, value=venta['periodo'])
                ws3.cell(row=idx+1, column=2, value=venta['cantidad'])
                ws3.cell(row=idx+1, column=3, value=f"Bs. {venta['efectivo']:.2f}")
                ws3.cell(row=idx+1, column=4, value=f"Bs. {venta['qr']:.2f}")
                ws3.cell(row=idx+1, column=5, value=f"Bs. {venta['credito']:.2f}")
                ws3.cell(row=idx+1, column=6, value=f"Bs. {venta['total']:.2f}")
            
            # ===== HOJA 4: VENTAS POR CATEGORÍA =====
            ws4 = wb.create_sheet("Ventas por Categoría")
            
            ventas_cat = self.obtener_ventas_por_categoria(fecha_desde, fecha_hasta)
            
            headers = ['Categoría', 'Cantidad Vendida', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws4.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="ffc107", end_color="ffc107", fill_type="solid")
            
            for idx, cat in enumerate(ventas_cat, 1):
                ws4.cell(row=idx+1, column=1, value=cat.get('categoria', 'Sin categoría'))
                ws4.cell(row=idx+1, column=2, value=cat['cantidad'])
                ws4.cell(row=idx+1, column=3, value=f"Bs. {cat['total']:.2f}")
            
            # Ajustar anchos de columna
            for ws in [ws1, ws2, ws3, ws4]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            os.makedirs('exports', exist_ok=True)
            filename = f"exports/estadisticas_{fecha_desde}_{fecha_hasta}.xlsx"
            wb.save(filename)
            
            print(f'✅ Reporte de estadísticas exportado: {filename}')
            return filename
            
        except Exception as e:
            print(f'❌ Error al exportar estadísticas: {e}')
            return None

# ============================================
# FIN DE LAS FUNCIONES DE ESTADÍSTICAS
# ============================================

# Ejecutar si se llama directamente
if __name__ == '__main__':
    db = Database()
    print('Base de datos lista para usar')

